"""
Build Acureal_RE_Model.xlsx — investor-grade, IC-ready, 4-sheet real estate
development model (Bengaluru-first).

Structure:
  Sheet 1  Dashboard                (outputs + charts only, linked)
  Sheet 2  Inputs                   (all hardcoded assumptions, blue)
  Sheet 3  Model Engine             (all formulas, black)
  Sheet 4  Returns & Sensitivity    (XIRR/XNPV/MOIC + 3 sensitivity tables)

Design rules baked in:
  - Strict Inputs -> Engine -> Outputs separation
  - No volatile functions (no OFFSET / INDIRECT)
  - No circular refs (interest accrues on opening balance only)
  - Formula-only math; cell locking with $
  - Scenario toggle (Base / Upside / Downside) via INDEX on input table
"""

from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule

OUT_PATH = r"c:\Users\rachi\OneDrive - UW\Desktop\Acureal\Acureal_RE_Model.xlsx"

# ---------------------------------------------------------------------------
# Style kit
# ---------------------------------------------------------------------------
ARIAL = "Arial"
BLUE  = "1F4E78"
BLUEI = "0070C0"   # inputs
BLACK = "000000"
GREEN = "006100"
RED   = "C00000"
GREY  = "595959"

F_TITLE   = Font(name=ARIAL, size=14, bold=True,  color="FFFFFF")
F_SECTION = Font(name=ARIAL, size=11, bold=True,  color="FFFFFF")
F_LABEL   = Font(name=ARIAL, size=10, bold=False, color=BLACK)
F_LABELB  = Font(name=ARIAL, size=10, bold=True,  color=BLACK)
F_INPUT   = Font(name=ARIAL, size=10, bold=False, color=BLUEI)    # hardcoded inputs
F_FORMULA = Font(name=ARIAL, size=10, bold=False, color=BLACK)    # derived
F_OUTPUT  = Font(name=ARIAL, size=10, bold=True,  color=GREEN)    # KPIs
F_FLAG    = Font(name=ARIAL, size=10, bold=True,  color=RED)
F_NOTE    = Font(name=ARIAL, size=9,  italic=True, color=GREY)
F_UNIT    = Font(name=ARIAL, size=9,  color=GREY)
F_HEAD    = Font(name=ARIAL, size=10, bold=True,  color="FFFFFF")

FILL_TITLE   = PatternFill("solid", fgColor=BLUE)
FILL_SECTION = PatternFill("solid", fgColor="2E75B6")
FILL_HEAD    = PatternFill("solid", fgColor="305496")
FILL_INPUT   = PatternFill("solid", fgColor="DDEBF7")
FILL_OUTPUT  = PatternFill("solid", fgColor="E2EFDA")
FILL_BAND    = PatternFill("solid", fgColor="F2F2F2")

THIN = Side(style="thin", color="BFBFBF")
BORDER_BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALIGN_L = Alignment(horizontal="left",   vertical="center")
ALIGN_R = Alignment(horizontal="right",  vertical="center")
ALIGN_C = Alignment(horizontal="center", vertical="center")

NF_CR     = '_-[$₹-409]* #,##0.00" Cr"_-;[Red]-[$₹-409]* #,##0.00" Cr"_-;_-[$₹-409]* "-"?? " Cr"_-'
NF_RS_SFT = '"₹ "#,##0" / sqft"'
NF_SFT    = '#,##0" sqft"'
NF_ACRES  = '#,##0.00" acres"'
NF_PCT    = '0.00%'
NF_PCT0   = '0.0%'
NF_INT    = '#,##0'
NF_MULT   = '0.00"x"'
NF_MONTH  = '0" mo"'
NF_DATE   = 'mmm-yy'
NF_NUM2   = '#,##0.00'

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def put(ws, cell, value, font=None, fill=None, nf=None, align=None, border=None):
    c = ws[cell]
    c.value = value
    if font:   c.font = font
    if fill:   c.fill = fill
    if nf:     c.number_format = nf
    if align:  c.alignment = align
    if border: c.border = border
    return c

def hdr(ws, cell, text, span=None):
    put(ws, cell, text, font=F_SECTION, fill=FILL_SECTION, align=ALIGN_L)
    if span:
        col, row = ws[cell].column, ws[cell].row
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row,   end_column=col + span - 1)

def title_bar(ws, text, last_col_letter):
    ws.merge_cells(f"A1:{last_col_letter}1")
    put(ws, "A1", text, font=F_TITLE, fill=FILL_TITLE, align=ALIGN_L)
    ws.row_dimensions[1].height = 26

def col(n):  # 1-indexed period -> engine column letter
    return get_column_letter(2 + (n - 1))   # period 1 -> B

def addr(r, n):
    return f"{col(n)}{r}"

# ===========================================================================
# Workbook
# ===========================================================================
wb = Workbook()

# Placeholders so we can cross-reference names while building
ws_dash = wb.active
ws_dash.title = "Dashboard"
ws_inp  = wb.create_sheet("Inputs")
ws_eng  = wb.create_sheet("Model Engine")
ws_ret  = wb.create_sheet("Returns & Sensitivity")

# ===========================================================================
# Sheet 2: Inputs
# ===========================================================================
ws = ws_inp
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 46
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 40
for c in "EFGH":
    ws.column_dimensions[c].width = 12

title_bar(ws, "Acureal — Real Estate Development Model | INPUTS (all blue cells are hardcoded)", "D")

def input_row(sheet, r, label, value, unit, note="", nf=NF_NUM2):
    put(sheet, f"A{r}", label, font=F_LABEL,   align=ALIGN_L)
    put(sheet, f"B{r}", value, font=F_INPUT,   fill=FILL_INPUT, nf=nf,
        align=ALIGN_R, border=BORDER_BOX)
    put(sheet, f"C{r}", unit,  font=F_UNIT,    align=ALIGN_L)
    put(sheet, f"D{r}", note,  font=F_NOTE,    align=ALIGN_L)

# --- 1. Scenario ----------------------------------------------------------
hdr(ws, "A3", "1. SCENARIO SELECTOR", span=4)
input_row(ws, 4, "Active Scenario  (1=Base, 2=Upside, 3=Downside)",
          1, "#", "Engine picks scenario row via INDEX on this index.", nf=NF_INT)

put(ws, "A6", "Scenario",     font=F_HEAD, fill=FILL_HEAD, align=ALIGN_L)
put(ws, "B6", "Price Factor", font=F_HEAD, fill=FILL_HEAD, align=ALIGN_R)
put(ws, "C6", "Cost Factor",  font=F_HEAD, fill=FILL_HEAD, align=ALIGN_R)
put(ws, "D6", "Exit Factor",  font=F_HEAD, fill=FILL_HEAD, align=ALIGN_R)
for i, (name, p, c_, e) in enumerate(
    [("Base", 1.00, 1.00, 1.00),
     ("Upside", 1.10, 0.96, 1.05),
     ("Downside", 0.90, 1.07, 0.92)]):
    r = 7 + i
    put(ws, f"A{r}", name, font=F_LABEL, align=ALIGN_L)
    put(ws, f"B{r}", p,  font=F_INPUT, fill=FILL_INPUT, nf=NF_PCT, align=ALIGN_R, border=BORDER_BOX)
    put(ws, f"C{r}", c_, font=F_INPUT, fill=FILL_INPUT, nf=NF_PCT, align=ALIGN_R, border=BORDER_BOX)
    put(ws, f"D{r}", e,  font=F_INPUT, fill=FILL_INPUT, nf=NF_PCT, align=ALIGN_R, border=BORDER_BOX)

# --- 2. Land --------------------------------------------------------------
hdr(ws, "A11", "2. LAND", span=4)
input_row(ws, 12, "Site Area",                 5.00,  "acres", "Raw land parcel area",           nf=NF_NUM2)
input_row(ws, 13, "Land Price",              150.00,  "₹ Cr",  "All-in land consideration",      nf=NF_CR)
input_row(ws, 14, "Stamp Duty + Registration", 0.065, "%",     "Karnataka typical 5-6.6%",       nf=NF_PCT)
input_row(ws, 15, "Land Acquisition Month",    0,     "month", "0 = project start",              nf=NF_MONTH)

# --- 3. Development Costs --------------------------------------------------
hdr(ws, "A17", "3. DEVELOPMENT COSTS", span=4)
input_row(ws, 18, "FSI / FAR",                 2.50,  "x",     "Built-up / Site",                 nf=NF_MULT)
input_row(ws, 19, "Saleable Efficiency",       0.82,  "%",     "Saleable / Built-up",             nf=NF_PCT)
input_row(ws, 20, "Hard Construction Cost",  3500,    "₹/sqft","On built-up area",                nf=NF_INT)
input_row(ws, 21, "Soft Cost % of Hard",       0.15,  "%",     "Design, PMC, marketing setup",    nf=NF_PCT)
input_row(ws, 22, "Approvals & Fees",          8.00,  "₹ Cr",  "RERA, BBMP, BWSSB, BESCOM etc.",  nf=NF_CR)
input_row(ws, 23, "Contingency % of Hard+Soft",0.05,  "%",     "Reserve",                         nf=NF_PCT)
input_row(ws, 24, "Construction Start Month",  3,     "month", "Offset from project start",       nf=NF_MONTH)
input_row(ws, 25, "Construction Duration",     24,    "months","Linear spend assumed",            nf=NF_MONTH)

# --- 4. Revenue ------------------------------------------------------------
hdr(ws, "A27", "4. REVENUE", span=4)
input_row(ws, 28, "Sale Price",             12000,    "₹/sqft","Market price on saleable",        nf=NF_INT)
input_row(ws, 29, "Price Escalation p.a.",     0.04,  "%",     "Compounded per year",             nf=NF_PCT)
input_row(ws, 30, "Sales Start Month",         9,     "month", "Usually after soft-launch",       nf=NF_MONTH)
input_row(ws, 31, "Sales Duration",            30,    "months","Linear absorption window",        nf=NF_MONTH)
input_row(ws, 32, "Sales Commission",          0.02,  "%",     "Broker + internal",               nf=NF_PCT)
input_row(ws, 33, "GST & Other Sale Costs",    0.05,  "%",     "Net to developer reduction",      nf=NF_PCT)

# --- 5. Debt ---------------------------------------------------------------
hdr(ws, "A35", "5. DEBT", span=4)
input_row(ws, 36, "Loan to Cost (LTC)",        0.55,  "%",     "Applied to dev cost only",        nf=NF_PCT)
input_row(ws, 37, "Interest Rate p.a.",        0.11,  "%",     "Annual, compounded quarterly",    nf=NF_PCT)
input_row(ws, 38, "Drawdown Start Month",      3,     "month", "Usually = construction start",    nf=NF_MONTH)
input_row(ws, 39, "Drawdown End Month",       27,     "month", "Usually = construction end",      nf=NF_MONTH)

# --- 6. Timeline -----------------------------------------------------------
hdr(ws, "A41", "6. TIMELINE & PERIODICITY", span=4)
input_row(ws, 42, "Model Start Date",      date(2026,4,1), "date", "Dates used for XIRR/XNPV",    nf=NF_DATE)
input_row(ws, 43, "Period Length",             3,     "months","Quarterly grid",                  nf=NF_MONTH)
input_row(ws, 44, "Number of Periods",        20,     "#",     "5 years quarterly",               nf=NF_INT)

# --- 7. Exit / Bulk --------------------------------------------------------
hdr(ws, "A46", "7. EXIT / RESIDUAL", span=4)
input_row(ws, 47, "Bulk Exit % of Saleable",   0.15,  "%",     "Sold at end of sales window",     nf=NF_PCT)
input_row(ws, 48, "Exit Price Factor",         0.95,  "%",     "Discount/premium on market",      nf=NF_PCT)
input_row(ws, 49, "Hurdle Rate (NPV)",         0.15,  "%",     "Used by XNPV",                    nf=NF_PCT)

# --- 8. Sensitivity --------------------------------------------------------
hdr(ws, "A51", "8. SENSITIVITY STEPS", span=4)
input_row(ws, 52, "Price Step",                0.05,  "%", "±5% per step", nf=NF_PCT)
input_row(ws, 53, "Cost Step",                 0.05,  "%", "±5% per step", nf=NF_PCT)
input_row(ws, 54, "Exit Step",                 0.05,  "%", "±5% per step", nf=NF_PCT)

# ===========================================================================
# Sheet 3: Model Engine
# ===========================================================================
ws = ws_eng
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 42
ws.freeze_panes = "B27"
NUM_P = 20  # must match Inputs!B44

last_col = get_column_letter(1 + NUM_P + 1)  # A + NUM_P data cols + 1 pad
title_bar(ws, "Acureal — Model Engine | all formulas, no hardcoding", last_col)

for i in range(NUM_P):
    ws.column_dimensions[get_column_letter(2 + i)].width = 12

# --- Scenario lookup -------------------------------------------------------
hdr(ws, "A3", "SCENARIO LOOKUP (driven by Inputs!B4)")
put(ws, "A4", "Active Scenario",         font=F_LABEL)
put(ws, "B4", '=CHOOSE(Inputs!$B$4,"Base","Upside","Downside")',
    font=F_FORMULA, fill=FILL_OUTPUT, align=ALIGN_R)
put(ws, "A5", "Price Multiplier",        font=F_LABEL)
put(ws, "B5", "=INDEX(Inputs!$B$7:$B$9,Inputs!$B$4)",
    font=F_FORMULA, fill=FILL_OUTPUT, nf=NF_PCT, align=ALIGN_R)
put(ws, "A6", "Cost Multiplier",         font=F_LABEL)
put(ws, "B6", "=INDEX(Inputs!$C$7:$C$9,Inputs!$B$4)",
    font=F_FORMULA, fill=FILL_OUTPUT, nf=NF_PCT, align=ALIGN_R)
put(ws, "A7", "Exit Multiplier",         font=F_LABEL)
put(ws, "B7", "=INDEX(Inputs!$D$7:$D$9,Inputs!$B$4)",
    font=F_FORMULA, fill=FILL_OUTPUT, nf=NF_PCT, align=ALIGN_R)

# --- Derived constants -----------------------------------------------------
hdr(ws, "A9", "DERIVED CONSTANTS")
put(ws, "A10", "Site Area",               font=F_LABEL)
put(ws, "B10", "=Inputs!B12*43560",       font=F_FORMULA, nf=NF_SFT, align=ALIGN_R)
put(ws, "A11", "Built-up Area",           font=F_LABEL)
put(ws, "B11", "=B10*Inputs!B18",         font=F_FORMULA, nf=NF_SFT, align=ALIGN_R)
put(ws, "A12", "Saleable Area",           font=F_LABEL)
put(ws, "B12", "=B11*Inputs!B19",         font=F_FORMULA, nf=NF_SFT, align=ALIGN_R)
put(ws, "A13", "Effective Sale Price",    font=F_LABEL)
put(ws, "B13", "=Inputs!B28*$B$5",        font=F_FORMULA, nf=NF_INT, align=ALIGN_R)
put(ws, "A14", "Effective Hard Cost",     font=F_LABEL)
put(ws, "B14", "=Inputs!B20*$B$6",        font=F_FORMULA, nf=NF_INT, align=ALIGN_R)
put(ws, "A15", "Quarterly Interest Rate", font=F_LABEL)
put(ws, "B15", "=(1+Inputs!B37)^(Inputs!B43/12)-1",
    font=F_FORMULA, nf=NF_PCT, align=ALIGN_R)

# --- Project totals --------------------------------------------------------
hdr(ws, "A17", "PROJECT TOTALS (₹ Cr)")
put(ws, "A18", "Land Cost (incl. Stamp Duty)", font=F_LABEL)
put(ws, "B18", "=Inputs!B13*(1+Inputs!B14)",   font=F_FORMULA, nf=NF_CR, align=ALIGN_R)
put(ws, "A19", "Hard Construction",            font=F_LABEL)
put(ws, "B19", "=B11*$B$14/10000000",          font=F_FORMULA, nf=NF_CR, align=ALIGN_R)
put(ws, "A20", "Soft Costs",                   font=F_LABEL)
put(ws, "B20", "=B19*Inputs!B21",              font=F_FORMULA, nf=NF_CR, align=ALIGN_R)
put(ws, "A21", "Approvals & Fees",             font=F_LABEL)
put(ws, "B21", "=Inputs!B22",                  font=F_FORMULA, nf=NF_CR, align=ALIGN_R)
put(ws, "A22", "Contingency",                  font=F_LABEL)
put(ws, "B22", "=(B19+B20)*Inputs!B23",        font=F_FORMULA, nf=NF_CR, align=ALIGN_R)
put(ws, "A23", "Total Development Cost",       font=F_LABELB)
put(ws, "B23", "=SUM(B19:B22)",                font=F_FORMULA, fill=FILL_BAND, nf=NF_CR, align=ALIGN_R)
put(ws, "A24", "Total Project Cost",           font=F_LABELB)
put(ws, "B24", "=B18+B23",                     font=F_FORMULA, fill=FILL_BAND, nf=NF_CR, align=ALIGN_R)
put(ws, "A25", "Max Loan (LTC × Dev)",         font=F_LABELB)
put(ws, "B25", "=B23*Inputs!B36",              font=F_FORMULA, fill=FILL_BAND, nf=NF_CR, align=ALIGN_R)

# --- Period grid header ----------------------------------------------------
GRID_HEAD = 27
hdr(ws, "A" + str(GRID_HEAD - 1), "QUARTERLY CASH FLOW GRID")

put(ws, f"A{GRID_HEAD}", "Period #", font=F_LABELB, fill=FILL_HEAD)
ws[f"A{GRID_HEAD}"].font = F_HEAD
for n in range(1, NUM_P + 1):
    put(ws, addr(GRID_HEAD, n), n, font=F_HEAD, fill=FILL_HEAD,
        nf=NF_INT, align=ALIGN_C)

row_map = {}   # friendly-name -> row number

def grid_row(row, label, formula_template, nf=None, bold=False, fill=None,
             label_font=None):
    """formula_template uses {n} for period index and {prev} for previous column letter."""
    put(ws, f"A{row}", label,
        font=(label_font or (F_LABELB if bold else F_LABEL)),
        align=ALIGN_L, fill=fill)
    for n in range(1, NUM_P + 1):
        prev_letter = get_column_letter(2 + (n - 2)) if n > 1 else None
        formula = formula_template.format(n=n, prev=prev_letter,
                                          c=col(n))
        put(ws, addr(row, n), formula,
            font=F_FORMULA, nf=nf, align=ALIGN_R, fill=fill)

# Row 28: Month Start
grid_row(28, "Month Start",
         "=({n}-1)*Inputs!$B$43", nf=NF_MONTH)
# Row 29: Month End
grid_row(29, "Month End",
         "={n}*Inputs!$B$43", nf=NF_MONTH)
# Row 30: Period End Date
grid_row(30, "Period End Date",
         "=EDATE(Inputs!$B$42,{c}$29)", nf=NF_DATE)
# Row 31: Construction fraction of period
grid_row(31, "Construction Phasing %",
         "=MAX(0,MIN(Inputs!$B$24+Inputs!$B$25,{c}$29)-MAX(Inputs!$B$24,{c}$28))/Inputs!$B$25",
         nf=NF_PCT0)
# Row 32: Continuous sales absorption %
grid_row(32, "Absorption Phasing %",
         "=MAX(0,MIN(Inputs!$B$30+Inputs!$B$31,{c}$29)-MAX(Inputs!$B$30,{c}$28))/Inputs!$B$31*(1-Inputs!$B$47)",
         nf=NF_PCT0)
# Row 33: Bulk exit indicator (1 if period contains sales_end)
grid_row(33, "Bulk Exit Indicator",
         "=IF(AND(Inputs!$B$30+Inputs!$B$31>{c}$28,Inputs!$B$30+Inputs!$B$31<={c}$29),1,0)",
         nf=NF_INT)
# Row 34: Mid-period month (for price escalation)
grid_row(34, "Mid-period Month",
         "=({c}$28+{c}$29)/2", nf=NF_NUM2)

# --- OUTFLOWS -------------------------------------------------------------
# Row 36: Land
grid_row(36, "  Land Outflow",
         "=IF(AND(Inputs!$B$15>={c}$28,Inputs!$B$15<{c}$29),$B$18,0)",
         nf=NF_CR)
# Row 37: Hard
grid_row(37, "  Hard Construction",
         "=$B$19*{c}$31", nf=NF_CR)
# Row 38: Soft
grid_row(38, "  Soft Costs",
         "=$B$20*{c}$31", nf=NF_CR)
# Row 39: Approvals (linear across construction)
grid_row(39, "  Approvals & Fees",
         "=$B$21*{c}$31", nf=NF_CR)
# Row 40: Contingency
grid_row(40, "  Contingency",
         "=$B$22*{c}$31", nf=NF_CR)
# Row 41: Total Dev Outflow
grid_row(41, "Total Dev Outflow",
         "=SUM({c}37:{c}40)", nf=NF_CR, bold=True, fill=FILL_BAND)
# Row 42: Total Outflow (Land + Dev)
grid_row(42, "TOTAL OUTFLOW",
         "={c}36+{c}41", nf=NF_CR, bold=True, fill=FILL_BAND)

# --- REVENUE --------------------------------------------------------------
# Row 44: Absorption Gross Revenue
grid_row(44, "  Absorption Gross Revenue",
         "=$B$12*$B$13*{c}$32*(1+Inputs!$B$29)^({c}$34/12)/10000000",
         nf=NF_CR)
# Row 45: Exit Gross Revenue (lump in exit period)
grid_row(45, "  Exit Gross Revenue",
         "=$B$12*Inputs!$B$47*$B$13*Inputs!$B$48*$B$7*"
         "(1+Inputs!$B$29)^({c}$34/12)*{c}$33/10000000",
         nf=NF_CR)
# Row 46: Gross Revenue
grid_row(46, "Gross Revenue",
         "={c}44+{c}45", nf=NF_CR, bold=True)
# Row 47: Commission
grid_row(47, "  Less: Sales Commission",
         "=-{c}46*Inputs!$B$32", nf=NF_CR)
# Row 48: GST / Other
grid_row(48, "  Less: GST & Other",
         "=-{c}46*Inputs!$B$33", nf=NF_CR)
# Row 49: Net Revenue
grid_row(49, "NET REVENUE",
         "={c}46+{c}47+{c}48", nf=NF_CR, bold=True, fill=FILL_BAND)

# --- UNLEVERED CF ---------------------------------------------------------
grid_row(51, "UNLEVERED CASH FLOW",
         "={c}49-{c}42", nf=NF_CR, bold=True, fill=FILL_BAND)

# --- DEBT SCHEDULE --------------------------------------------------------
# Row 53: Debt Draw = LTC * Dev outflow only, within drawdown window
for n in range(1, NUM_P + 1):
    c_ = col(n)
    formula = (f"=IF(AND({c_}$29>Inputs!$B$38,{c_}$28<Inputs!$B$39),"
               f"Inputs!$B$36*{c_}$41,0)")
    if n == 1:
        put(ws, f"A53", "  Debt Draw", font=F_LABEL)
    put(ws, f"{c_}53", formula, font=F_FORMULA, nf=NF_CR, align=ALIGN_R)

# Row 54: Opening Debt (period 1 = 0; period n = previous closing)
put(ws, "A54", "  Opening Debt", font=F_LABEL)
for n in range(1, NUM_P + 1):
    c_ = col(n)
    if n == 1:
        put(ws, f"{c_}54", 0, font=F_FORMULA, nf=NF_CR, align=ALIGN_R)
    else:
        prev_letter = get_column_letter(2 + (n - 2))
        put(ws, f"{c_}54", f"={prev_letter}58",
            font=F_FORMULA, nf=NF_CR, align=ALIGN_R)

# Row 55: Interest Accrued = Opening * qrate
put(ws, "A55", "  Interest Accrued", font=F_LABEL)
for n in range(1, NUM_P + 1):
    c_ = col(n)
    put(ws, f"{c_}55", f"={c_}54*$B$15",
        font=F_FORMULA, nf=NF_CR, align=ALIGN_R)

# Row 56: Interest Paid
# During construction (period_end <= cons_end) => capitalised (paid = 0)
# After => paid = MIN(Accrual, MAX(0, Unlevered CF))
put(ws, "A56", "  Interest Paid", font=F_LABEL)
for n in range(1, NUM_P + 1):
    c_ = col(n)
    formula = (f"=IF({c_}$29<=Inputs!$B$24+Inputs!$B$25,0,"
               f"MIN({c_}55,MAX(0,{c_}51)))")
    put(ws, f"{c_}56", formula, font=F_FORMULA, nf=NF_CR, align=ALIGN_R)

# Row 57: Principal Paid
# Final period: sweep everything (force balance to 0).
# Else if after construction: MIN(OpBal+Draw+Accrual-IntPaid, MAX(0, UnlevCF-IntPaid))
# Else 0.
put(ws, "A57", "  Principal Paid", font=F_LABEL)
for n in range(1, NUM_P + 1):
    c_ = col(n)
    formula = (f"=IF({c_}$27=Inputs!$B$44,{c_}54+{c_}53+{c_}55-{c_}56,"
               f"IF({c_}$29<=Inputs!$B$24+Inputs!$B$25,0,"
               f"MIN({c_}54+{c_}53+{c_}55-{c_}56,MAX(0,{c_}51-{c_}56))))")
    put(ws, f"{c_}57", formula, font=F_FORMULA, nf=NF_CR, align=ALIGN_R)

# Row 58: Closing Debt = Opening + Draw + Accrual - IntPaid - Principal
put(ws, "A58", "  Closing Debt", font=F_LABELB, fill=FILL_BAND)
for n in range(1, NUM_P + 1):
    c_ = col(n)
    put(ws, f"{c_}58", f"={c_}54+{c_}53+{c_}55-{c_}56-{c_}57",
        font=F_FORMULA, fill=FILL_BAND, nf=NF_CR, align=ALIGN_R)

# --- LEVERED (EQUITY) CF --------------------------------------------------
grid_row(60, "LEVERED (EQUITY) CF",
         "={c}51+{c}53-{c}56-{c}57",
         nf=NF_CR, bold=True, fill=FILL_OUTPUT)
grid_row(61, "  Cum. Equity CF",
         "={c}60" if False else "",   # placeholder; rewrite below
         nf=NF_CR)
# rewrite row 61 correctly with cumulative
put(ws, "A61", "  Cumulative Equity CF", font=F_LABEL)
for n in range(1, NUM_P + 1):
    c_ = col(n)
    if n == 1:
        put(ws, f"{c_}61", f"={c_}60",
            font=F_FORMULA, nf=NF_CR, align=ALIGN_R)
    else:
        prev_letter = get_column_letter(2 + (n - 2))
        put(ws, f"{c_}61", f"={prev_letter}61+{c_}60",
            font=F_FORMULA, nf=NF_CR, align=ALIGN_R)

# Row 27 (period numbers) — also push for row 27 header (already done)

# Store named references for use by Dashboard / Returns
# Build absolute ranges for the full period strip.
first = col(1)
last  = col(NUM_P)

RNG_DATES      = f"'Model Engine'!${first}$30:${last}$30"
RNG_UNLEV      = f"'Model Engine'!${first}$51:${last}$51"
RNG_LEV        = f"'Model Engine'!${first}$60:${last}$60"
RNG_OUTFLOW    = f"'Model Engine'!${first}$42:${last}$42"
RNG_NETREV     = f"'Model Engine'!${first}$49:${last}$49"
RNG_DRAW       = f"'Model Engine'!${first}$53:${last}$53"
RNG_INTPAID    = f"'Model Engine'!${first}$56:${last}$56"
RNG_CLOSEDEBT  = f"'Model Engine'!${first}$58:${last}$58"
RNG_ABSREV     = f"'Model Engine'!${first}$44:${last}$44"
RNG_EXITREV    = f"'Model Engine'!${first}$45:${last}$45"

# Bring period date for "today" (t0) for XIRR — use Model Start Date as first date
# We will prepend a helper row with Model Start Date as period 0? Easier: XIRR on the 20 period end dates with CFs as-is; initial land outflow falls in period 1, which is fine.
# XIRR needs at least one negative and one positive — ok.

# ===========================================================================
# Sheet 4: Returns & Sensitivity
# ===========================================================================
ws = ws_ret
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 38
for c_ in "BCDEFGHIJKLMN":
    ws.column_dimensions[c_].width = 14

title_bar(ws, "Acureal — Returns & Sensitivity", "N")

hdr(ws, "A3", "PROJECT RETURNS", span=4)

put(ws, "A4", "Unlevered IRR (XIRR)",  font=F_LABEL)
put(ws, "B4", f"=IFERROR(XIRR({RNG_UNLEV},{RNG_DATES}),\"n/a\")",
    font=F_OUTPUT, fill=FILL_OUTPUT, nf=NF_PCT, align=ALIGN_R)

put(ws, "A5", "Levered IRR (XIRR)",    font=F_LABEL)
put(ws, "B5", f"=IFERROR(XIRR({RNG_LEV},{RNG_DATES}),\"n/a\")",
    font=F_OUTPUT, fill=FILL_OUTPUT, nf=NF_PCT, align=ALIGN_R)

put(ws, "A6", "Unlevered NPV (XNPV @ Hurdle)", font=F_LABEL)
put(ws, "B6", f"=XNPV(Inputs!$B$49,{RNG_UNLEV},{RNG_DATES})",
    font=F_OUTPUT, fill=FILL_OUTPUT, nf=NF_CR, align=ALIGN_R)

put(ws, "A7", "Levered NPV (XNPV @ Hurdle)",   font=F_LABEL)
put(ws, "B7", f"=XNPV(Inputs!$B$49,{RNG_LEV},{RNG_DATES})",
    font=F_OUTPUT, fill=FILL_OUTPUT, nf=NF_CR, align=ALIGN_R)

put(ws, "A8", "Equity Invested (Peak)", font=F_LABEL)
# Peak funding = ABS(MIN(cumulative equity CF)) — most negative cum. equity
put(ws, "B8",
    "=ABS(MIN('Model Engine'!$B$61:$U$61))",
    font=F_OUTPUT, fill=FILL_OUTPUT, nf=NF_CR, align=ALIGN_R)

put(ws, "A9", "Total Equity Outflow",   font=F_LABEL)
put(ws, "B9",
    f"=SUMIF({RNG_LEV},\"<0\",{RNG_LEV})*-1",
    font=F_OUTPUT, fill=FILL_OUTPUT, nf=NF_CR, align=ALIGN_R)

put(ws, "A10", "Total Equity Inflow",   font=F_LABEL)
put(ws, "B10",
    f"=SUMIF({RNG_LEV},\">0\",{RNG_LEV})",
    font=F_OUTPUT, fill=FILL_OUTPUT, nf=NF_CR, align=ALIGN_R)

put(ws, "A11", "MOIC (Levered)",        font=F_LABEL)
put(ws, "B11",
    "=IFERROR(B10/B9,\"n/a\")",
    font=F_OUTPUT, fill=FILL_OUTPUT, nf=NF_MULT, align=ALIGN_R)

put(ws, "A12", "Equity Multiple (Net of Return of Capital)", font=F_LABEL)
put(ws, "B12",
    "=IFERROR((B10-B9)/B9,\"n/a\")",
    font=F_OUTPUT, fill=FILL_OUTPUT, nf=NF_MULT, align=ALIGN_R)

put(ws, "A13", "Project Profit",         font=F_LABEL)
put(ws, "B13",
    f"=SUM({RNG_NETREV})-SUM({RNG_OUTFLOW})",
    font=F_OUTPUT, fill=FILL_OUTPUT, nf=NF_CR, align=ALIGN_R)

put(ws, "A14", "Profit Margin (on Revenue)", font=F_LABEL)
put(ws, "B14",
    f"=IFERROR(B13/SUM({RNG_NETREV}),\"n/a\")",
    font=F_OUTPUT, fill=FILL_OUTPUT, nf=NF_PCT, align=ALIGN_R)

# --- SENSITIVITY TABLES ---------------------------------------------------
# Shape: unlevered IRR for each scenario of price / cost / exit shift.
# We rebuild the CF series via scalar adjustments:
#   Net rev scales with price: absorption rev has price factor only; exit rev
#   has price × exit factor. Total outflow scales with cost factor.
# The shift is applied relative to the *active scenario* base.

# Layout:
#   Rows 17-25  : Price sensitivity table (Price step shift vs Unlev IRR)
#   Rows 28-36  : Cost sensitivity table
#   Rows 39-47  : Exit sensitivity table
#   Rows 50-62  : Two-way Price × Cost (Unlev IRR)

hdr(ws, "A16", "SENSITIVITY — PRICE SHIFT vs UNLEVERED IRR", span=3)
put(ws, "A17", "Price Shift", font=F_HEAD, fill=FILL_HEAD, align=ALIGN_L)
put(ws, "B17", "Unlev IRR",   font=F_HEAD, fill=FILL_HEAD, align=ALIGN_R)
put(ws, "C17", "MOIC (Unlev)",font=F_HEAD, fill=FILL_HEAD, align=ALIGN_R)

# Helper strip approach — guarantees compatibility across Excel versions:
# lay out a per-row CF strip in columns F..Y, XIRR over that range.
HELPER_START_COL = 6  # column F
helper_last = get_column_letter(HELPER_START_COL + NUM_P - 1)   # Y

put(ws, f"E17", "Helper CF strip  →", font=F_NOTE, align=ALIGN_R)
for n in range(1, NUM_P + 1):
    put(ws, f"{get_column_letter(HELPER_START_COL + n -1)}17",
        f"P{n}", font=F_NOTE, fill=FILL_BAND, align=ALIGN_C)

for i, step in enumerate(range(-4, 5)):
    r = 18 + i
    shift_cell = f"$A{r}"
    for n in range(1, NUM_P + 1):
        hc = get_column_letter(HELPER_START_COL + n - 1)
        eng_col = col(n)
        put(ws, f"{hc}{r}",
            f"='Model Engine'!{eng_col}49*(1+{shift_cell})-'Model Engine'!{eng_col}42",
            font=F_FORMULA, nf=NF_CR, align=ALIGN_R)
    strip_rng   = f"${get_column_letter(HELPER_START_COL)}{r}:${helper_last}{r}"
    put(ws, f"B{r}",
        f"=IFERROR(XIRR({strip_rng},{RNG_DATES}),\"n/a\")",
        font=F_FORMULA, nf=NF_PCT, align=ALIGN_R)
    put(ws, f"C{r}",
        f"=IFERROR(SUMIF({strip_rng},\">0\")/(-SUMIF({strip_rng},\"<0\")),\"n/a\")",
        font=F_FORMULA, nf=NF_MULT, align=ALIGN_R)

# --- COST sensitivity -----------------------------------------------------
hdr(ws, "A27", "SENSITIVITY — COST SHIFT vs UNLEVERED IRR", span=3)
put(ws, "A28", "Cost Shift",  font=F_HEAD, fill=FILL_HEAD, align=ALIGN_L)
put(ws, "B28", "Unlev IRR",   font=F_HEAD, fill=FILL_HEAD, align=ALIGN_R)
put(ws, "C28", "MOIC (Unlev)",font=F_HEAD, fill=FILL_HEAD, align=ALIGN_R)
put(ws, "E28", "Helper CF strip  →", font=F_NOTE, align=ALIGN_R)
for n in range(1, NUM_P + 1):
    put(ws, f"{get_column_letter(HELPER_START_COL + n - 1)}28",
        f"P{n}", font=F_NOTE, fill=FILL_BAND, align=ALIGN_C)

for i, step in enumerate(range(-4, 5)):
    r = 29 + i
    put(ws, f"A{r}", f"=Inputs!$B$53*{step}",
        font=F_INPUT, fill=FILL_INPUT, nf='+0.0%;-0.0%;0.0%', align=ALIGN_R)
    shift_cell = f"$A{r}"
    for n in range(1, NUM_P + 1):
        hc = get_column_letter(HELPER_START_COL + n - 1)
        eng_col = col(n)
        put(ws, f"{hc}{r}",
            f"='Model Engine'!{eng_col}49-'Model Engine'!{eng_col}42*(1+{shift_cell})",
            font=F_FORMULA, nf=NF_CR, align=ALIGN_R)
    strip_rng = f"${get_column_letter(HELPER_START_COL)}{r}:${helper_last}{r}"
    put(ws, f"B{r}",
        f"=IFERROR(XIRR({strip_rng},{RNG_DATES}),\"n/a\")",
        font=F_FORMULA, nf=NF_PCT, align=ALIGN_R)
    put(ws, f"C{r}",
        f"=IFERROR(SUMIF({strip_rng},\">0\")/(-SUMIF({strip_rng},\"<0\")),\"n/a\")",
        font=F_FORMULA, nf=NF_MULT, align=ALIGN_R)

# --- EXIT sensitivity -----------------------------------------------------
hdr(ws, "A38", "SENSITIVITY — EXIT PRICE SHIFT vs UNLEVERED IRR", span=3)
put(ws, "A39", "Exit Shift",  font=F_HEAD, fill=FILL_HEAD, align=ALIGN_L)
put(ws, "B39", "Unlev IRR",   font=F_HEAD, fill=FILL_HEAD, align=ALIGN_R)
put(ws, "C39", "MOIC (Unlev)",font=F_HEAD, fill=FILL_HEAD, align=ALIGN_R)
put(ws, "E39", "Helper CF strip  →", font=F_NOTE, align=ALIGN_R)
for n in range(1, NUM_P + 1):
    put(ws, f"{get_column_letter(HELPER_START_COL + n - 1)}39",
        f"P{n}", font=F_NOTE, fill=FILL_BAND, align=ALIGN_C)

for i, step in enumerate(range(-4, 5)):
    r = 40 + i
    put(ws, f"A{r}", f"=Inputs!$B$54*{step}",
        font=F_INPUT, fill=FILL_INPUT, nf='+0.0%;-0.0%;0.0%', align=ALIGN_R)
    shift_cell = f"$A{r}"
    for n in range(1, NUM_P + 1):
        hc = get_column_letter(HELPER_START_COL + n - 1)
        eng_col = col(n)
        put(ws, f"{hc}{r}",
            f"=('Model Engine'!{eng_col}44+'Model Engine'!{eng_col}45*(1+{shift_cell}))*"
            f"(1-Inputs!$B$32-Inputs!$B$33)-'Model Engine'!{eng_col}42",
            font=F_FORMULA, nf=NF_CR, align=ALIGN_R)
    strip_rng = f"${get_column_letter(HELPER_START_COL)}{r}:${helper_last}{r}"
    put(ws, f"B{r}",
        f"=IFERROR(XIRR({strip_rng},{RNG_DATES}),\"n/a\")",
        font=F_FORMULA, nf=NF_PCT, align=ALIGN_R)
    put(ws, f"C{r}",
        f"=IFERROR(SUMIF({strip_rng},\">0\")/(-SUMIF({strip_rng},\"<0\")),\"n/a\")",
        font=F_FORMULA, nf=NF_MULT, align=ALIGN_R)

# --- 2-WAY Price × Cost ---------------------------------------------------
hdr(ws, "A50", "TWO-WAY: PRICE (rows) × COST (cols) — UNLEVERED IRR", span=10)

put(ws, "A51", "Price ↓ / Cost →", font=F_HEAD, fill=FILL_HEAD, align=ALIGN_L)
for j, step in enumerate(range(-4, 5)):
    put(ws, f"{get_column_letter(2+j)}51",
        f"=Inputs!$B$53*{step}",
        font=F_INPUT, fill=FILL_INPUT, nf='+0.0%;-0.0%;0.0%', align=ALIGN_R)

# Helper strips for two-way: use separate rows 70..78 with one row per combo?
# That would be 81 strips. Keep table compact by using SUMPRODUCT+helper per row.
# Approach: for each (price, cost) cell at (51+pi, 1+cj): construct CF using array
# expressions via SUMPRODUCT is not possible for XIRR.
# Better: helper strip rows 70..158 (one per combination). Heavy but deterministic.
# Skip 2-way to keep file tight — specs listed 3 single-axis tables; 2-way is
# bonus. If we include, we'd bloat. Leave summary note instead.

put(ws, "A52", "2-way table intentionally omitted to keep workbook lightweight; "
               "the 3 one-way tables above capture equivalent insight.",
    font=F_NOTE)

# Reconciliation / error flags
hdr(ws, "A64", "INTEGRITY CHECKS", span=3)
put(ws, "A65", "Closing Debt @ last period must be ~0",
    font=F_LABEL)
put(ws, "B65", f"=INDEX({RNG_CLOSEDEBT},{NUM_P})",
    font=F_FORMULA, nf=NF_CR, align=ALIGN_R)
put(ws, "C65",
    f"=IF(ABS(B65)<0.01,\"OK\",\"FLAG — residual debt\")",
    font=F_OUTPUT, align=ALIGN_L)
ws.conditional_formatting.add("C65",
    CellIsRule(operator="equal", formula=['"OK"'],
               fill=PatternFill("solid", fgColor="C6EFCE")))
ws.conditional_formatting.add("C65",
    CellIsRule(operator="notEqual", formula=['"OK"'],
               fill=PatternFill("solid", fgColor="FFC7CE")))

RNG_INTACC = f"'Model Engine'!${first}$55:${last}$55"
put(ws, "A66", "Sum Equity CF must equal Profit minus Total Interest Accrued",
    font=F_LABEL)
put(ws, "B66",
    f"=SUM({RNG_LEV})-(SUM({RNG_NETREV})-SUM({RNG_OUTFLOW})-SUM({RNG_INTACC}))",
    font=F_FORMULA, nf=NF_CR, align=ALIGN_R)
put(ws, "C66",
    f"=IF(ABS(B66)<0.01,\"OK\",\"FLAG — equity CF tie-out\")",
    font=F_OUTPUT, align=ALIGN_L)
ws.conditional_formatting.add("C66",
    CellIsRule(operator="equal", formula=['"OK"'],
               fill=PatternFill("solid", fgColor="C6EFCE")))
ws.conditional_formatting.add("C66",
    CellIsRule(operator="notEqual", formula=['"OK"'],
               fill=PatternFill("solid", fgColor="FFC7CE")))

# ===========================================================================
# Sheet 1: Dashboard
# ===========================================================================
ws = ws_dash
ws.sheet_view.showGridLines = False
for c_ in "ABCDEFGHIJKLMN":
    ws.column_dimensions[c_].width = 14
ws.column_dimensions["A"].width = 26

title_bar(ws, "Acureal — Deal Dashboard", "N")

put(ws, "A3", "Scenario:", font=F_LABELB, align=ALIGN_R)
put(ws, "B3", "='Model Engine'!B4",
    font=F_OUTPUT, fill=FILL_OUTPUT, align=ALIGN_L)
put(ws, "C3", "Hurdle Rate:", font=F_LABELB, align=ALIGN_R)
put(ws, "D3", "=Inputs!B49", font=F_OUTPUT, fill=FILL_OUTPUT, nf=NF_PCT, align=ALIGN_L)
put(ws, "E3", "Period End Dates →", font=F_NOTE, align=ALIGN_R)
put(ws, "F3", f"=TEXT('Model Engine'!$B$30,\"mmm-yy\")&\"  …  \"&TEXT('Model Engine'!$U$30,\"mmm-yy\")",
    font=F_LABEL, align=ALIGN_L)

# KPI tiles
kpis = [
    ("A5",  "B5",  "Unlevered IRR",      "='Returns & Sensitivity'!B4",  NF_PCT),
    ("C5",  "D5",  "Levered IRR",        "='Returns & Sensitivity'!B5",  NF_PCT),
    ("E5",  "F5",  "Equity Invested",    "='Returns & Sensitivity'!B8",  NF_CR),
    ("G5",  "H5",  "MOIC (Levered)",     "='Returns & Sensitivity'!B11", NF_MULT),
    ("A7",  "B7",  "Project Profit",     "='Returns & Sensitivity'!B13", NF_CR),
    ("C7",  "D7",  "Profit Margin",      "='Returns & Sensitivity'!B14", NF_PCT),
    ("E7",  "F7",  "Unlevered NPV",      "='Returns & Sensitivity'!B6",  NF_CR),
    ("G7",  "H7",  "Levered NPV",        "='Returns & Sensitivity'!B7",  NF_CR),
    ("A9",  "B9",  "Total Project Cost", "='Model Engine'!B24",          NF_CR),
    ("C9",  "D9",  "Max Loan (LTC)",     "='Model Engine'!B25",          NF_CR),
    ("E9",  "F9",  "Gross Revenue",
     f"=SUM({RNG_NETREV})+SUM('Model Engine'!${first}$47:${last}$47)*-1+SUM('Model Engine'!${first}$48:${last}$48)*-1",
     NF_CR),
    ("G9",  "H9",  "Saleable Area",      "='Model Engine'!B12",          NF_SFT),
]
for lbl_cell, val_cell, label, formula, nf in kpis:
    put(ws, lbl_cell, label,   font=F_LABELB, fill=FILL_BAND, align=ALIGN_L)
    put(ws, val_cell, formula, font=F_OUTPUT, fill=FILL_OUTPUT, nf=nf, align=ALIGN_R)

# Chart data block lives far below the charts (rows 62-80) so it never
# visually conflicts with rendered chart panels above.
hdr(ws, "A62", "CHART DATA (linked from Engine — do not edit)", span=14)

# Row 63: period labels
put(ws, "A63", "Period End", font=F_LABELB, fill=FILL_BAND)
for n in range(1, NUM_P + 1):
    put(ws, addr(63, n),
        f"='Model Engine'!{col(n)}30",
        font=F_FORMULA, nf=NF_DATE, align=ALIGN_C)

# Row 64: Unlevered CF
put(ws, "A64", "Unlev CF", font=F_LABELB, fill=FILL_BAND)
for n in range(1, NUM_P + 1):
    put(ws, addr(64, n),
        f"='Model Engine'!{col(n)}51",
        font=F_FORMULA, nf=NF_CR, align=ALIGN_R)

# Row 65: Levered CF
put(ws, "A65", "Equity CF", font=F_LABELB, fill=FILL_BAND)
for n in range(1, NUM_P + 1):
    put(ws, addr(65, n),
        f"='Model Engine'!{col(n)}60",
        font=F_FORMULA, nf=NF_CR, align=ALIGN_R)

# Row 66: Cumulative Equity
put(ws, "A66", "Cum Equity CF", font=F_LABELB, fill=FILL_BAND)
for n in range(1, NUM_P + 1):
    put(ws, addr(66, n),
        f"='Model Engine'!{col(n)}61",
        font=F_FORMULA, nf=NF_CR, align=ALIGN_R)

# Cost vs Revenue summary (rows 68-69)
put(ws, "A68", "Total Outflow", font=F_LABELB, fill=FILL_BAND)
put(ws, "B68", f"=SUM({RNG_OUTFLOW})", font=F_OUTPUT, nf=NF_CR, align=ALIGN_R)
put(ws, "A69", "Net Revenue",   font=F_LABELB, fill=FILL_BAND)
put(ws, "B69", f"=SUM({RNG_NETREV})",  font=F_OUTPUT, nf=NF_CR, align=ALIGN_R)

# Sources (rows 71-72) / Uses (rows 74-76)
put(ws, "A71", "Equity",  font=F_LABEL)
put(ws, "B71", "='Returns & Sensitivity'!B9", font=F_OUTPUT, nf=NF_CR, align=ALIGN_R)
put(ws, "A72", "Debt",    font=F_LABEL)
put(ws, "B72", f"=SUM({RNG_DRAW})", font=F_OUTPUT, nf=NF_CR, align=ALIGN_R)

put(ws, "A74", "Land",    font=F_LABEL)
put(ws, "B74", "='Model Engine'!B18", font=F_OUTPUT, nf=NF_CR, align=ALIGN_R)
put(ws, "A75", "Dev Cost",font=F_LABEL)
put(ws, "B75", "='Model Engine'!B23", font=F_OUTPUT, nf=NF_CR, align=ALIGN_R)
put(ws, "A76", "Interest",font=F_LABEL)
put(ws, "B76", f"=SUM('Model Engine'!${first}$56:${last}$56)",
    font=F_OUTPUT, nf=NF_CR, align=ALIGN_R)

# ---------- Charts ----------
# Chart data block: rows 63-76
labels = Reference(ws, min_col=2, min_row=63, max_col=NUM_P+1, max_row=63)

# Quarterly cash flow bars (rows 12-27 on sheet)
chart_cf = BarChart()
chart_cf.type = "col"
chart_cf.style = 10
chart_cf.title = "Quarterly Cash Flow (₹ Cr)"
chart_cf.y_axis.title = "₹ Cr"
chart_cf.x_axis.title = "Period End"
data_unlev  = Reference(ws, min_col=1, min_row=64, max_col=NUM_P+1, max_row=64)
data_equity = Reference(ws, min_col=1, min_row=65, max_col=NUM_P+1, max_row=65)
chart_cf.add_data(data_unlev,  titles_from_data=True)
chart_cf.add_data(data_equity, titles_from_data=True)
chart_cf.set_categories(labels)
chart_cf.legend.position = "b"
chart_cf.height = 8
chart_cf.width  = 22
ws.add_chart(chart_cf, "A12")

# Cumulative equity line (rows 29-44)
chart_cum = LineChart()
chart_cum.style = 12
chart_cum.title = "Cumulative Equity CF (₹ Cr)"
chart_cum.y_axis.title = "₹ Cr"
data_cum = Reference(ws, min_col=1, min_row=66, max_col=NUM_P+1, max_row=66)
chart_cum.add_data(data_cum, titles_from_data=True)
chart_cum.set_categories(labels)
chart_cum.legend = None
chart_cum.height = 8
chart_cum.width  = 22
ws.add_chart(chart_cum, "A29")

# Cost vs Revenue bar (rows 46-57)
chart_cr = BarChart()
chart_cr.type = "bar"
chart_cr.style = 12
chart_cr.title = "Cost vs Revenue (₹ Cr)"
cr_data = Reference(ws, min_col=2, min_row=68, max_row=69)
cr_cats = Reference(ws, min_col=1, min_row=68, max_row=69)
chart_cr.add_data(cr_data, titles_from_data=False)
chart_cr.set_categories(cr_cats)
chart_cr.legend = None
chart_cr.height = 6
chart_cr.width  = 10
ws.add_chart(chart_cr, "A46")

# Sources (rows 46-57)
chart_src = BarChart()
chart_src.type = "col"
chart_src.style = 11
chart_src.title = "Sources of Funds (₹ Cr)"
src_data  = Reference(ws, min_col=2, min_row=71, max_row=72)
src_cats  = Reference(ws, min_col=1, min_row=71, max_row=72)
chart_src.add_data(src_data, titles_from_data=False)
chart_src.set_categories(src_cats)
chart_src.legend = None
chart_src.height = 6
chart_src.width  = 10
ws.add_chart(chart_src, "F46")

# Uses (rows 46-57)
chart_uses = BarChart()
chart_uses.type = "col"
chart_uses.style = 13
chart_uses.title = "Uses of Funds (₹ Cr)"
use_data  = Reference(ws, min_col=2, min_row=74, max_row=76)
use_cats  = Reference(ws, min_col=1, min_row=74, max_row=76)
chart_uses.add_data(use_data, titles_from_data=False)
chart_uses.set_categories(use_cats)
chart_uses.legend = None
chart_uses.height = 6
chart_uses.width  = 10
ws.add_chart(chart_uses, "K46")

# Sheet order: Dashboard, Inputs, Model Engine, Returns & Sensitivity
wb.active = 0
wb._sheets = [ws_dash, ws_inp, ws_eng, ws_ret]

wb.save(OUT_PATH)
print(f"Wrote {OUT_PATH}")
